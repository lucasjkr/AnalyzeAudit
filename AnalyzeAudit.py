import argparse, csv, json, msal, logging, os, time
import requests_cache
from requests_cache import DO_NOT_CACHE, NEVER_EXPIRE, EXPIRE_IMMEDIATELY, CachedSession
from datetime import timedelta
from datetime import datetime
from dotenv import dotenv_values
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
import urllib.parse

class AnalyzeMicrosoftAuditLog():
    input: str

    def __init__(self):
        # input file
        self.input = ""

        # load values for API keys from .env file
        self.config = dotenv_values()

        # ip_ignore_list .env entry to JSON list. If entry doesn't exist in .env file, then create an empty list
        self.ip_ignore_list = json.loads( self.config.get('IP_IGNORE_LIST', "[]") )

        self.counter = {}
        self.start_time = time.time()

        # Create empty workbook:
        self.workbook = Workbook()

        # cache graph requests indefinitely. This is because once we've fetched a mails metadata with its
        # InternetMessageId, that metadata will never change, so there's no point in letting it expire.
        # Also, makes re-running the script MUCH faster. Docs: https://pypi.org/project/requests-cache/
        self.session = requests_cache.CachedSession('requests_cache', expire_after=NEVER_EXPIRE)

        # bearer token for Graph API - make sure the first token is already expired
        self.token_expires_at = datetime.now() + timedelta(hours=-1)
        self.token = ""

    def __exit__(self):
        # this function empty for now, placeholder for if new logic added
        pass

    def write_to_worksheet(self, sheet, data):
        # If worksheet doesn't exist, create worksheet and write header row, then proceed to rest of function.
        if sheet not in self.workbook.sheetnames:
            worksheet = self.workbook.create_sheet(title=sheet)
            worksheet.append(list(data.keys()))

        # now insert the row of values that need to be inserted (action happens whether header row was created or not
        worksheet = self.workbook[sheet]
        worksheet.append(list(data.values()))

    def get_bearer_token(self):
        # code for retrieving a Graph token, slightly adapted from Microsofts example code
        # this should not be called directly, instead call bearer_token() below, which caches the token
        # rather than request a new token with each API call.
        app = msal.ConfidentialClientApplication(
            self.config['CLIENT_ID'],
            authority="https://login.microsoftonline.com/" + self.config['TENANT_ID'],
            client_credential=self.config['SECRET'],
        )
        result = app.acquire_token_silent(["https://graph.microsoft.com/.default"], account=None)
        if not result:
            logging.info("No suitable token exists in cache. Let's get a new one from AAD.")
            result = app.acquire_token_for_client(scopes=["https://graph.microsoft.com/.default"])

        if "access_token" in result:
            logging.info("Authentication succeeded. Token acquired")
            # return result["access_token"]
            return result
        else:
            logging.critical("Authentication failed: " + result.get("error_description", "No error description available"))
            raise Exception(
                "Authentication failed: " + result.get("error_description", "No error description available"))

    # cache the Graph token for 100 seconds LESS than the actual expiration time
    def bearer_token(self):
        if datetime.now() > self.token_expires_at:
            token_response = self.get_bearer_token()
            self.token = token_response["access_token"]
            self.token_expires_at = datetime.today() + timedelta(seconds=(token_response["expires_in"] - 100))
            return self.token
        else:
            return self.token

    # Scans an audit export and displays the operations occurring with in. Useful to determine whether new functions,
    # parsers or outputs need to be coded.
    def show_operations(self):
        with open(self.input) as csv_file:
            operations = []
            csv_reader = csv.DictReader(csv_file)
            for row in csv_reader:
                if row['Operation'] not in operations:
                    operations.append(row['Operation'])
                    # print(f"{row['Operation']} * {json.dumps(row['AuditData'])}")
        for op in sorted(operations):
            print(op)
        exit()

    def increase_counter(self, item):
        if item in self.counter:
            self.counter[item] += 1
        else:
            self.counter[item] = 1

    # Looks up each message from the Graph API in order to obtain metadata to assist with the review.
    def get_message(self, user, internet_message_id):
        internet_message_id = urllib.parse.quote(internet_message_id)
        headers = {
            'Content-Type': 'application/json',
            'Authorization': 'Bearer ' + self.bearer_token()
        }
        response = self.session.get(
            f"https://graph.microsoft.com/v1.0/users/{user}/messages?$filter=internetMessageId eq '{internet_message_id}'",
            headers=headers)
        try:
            return response.json()
        except Exception:
            return {}

    def analyze_update_inbox_rule(self, audit_data):
        # if you make any changes to the structure of the export field which is written to Excel,
        # those changes must also be made to "analyze_mail_rule" function below
        export = {
            'datetime': audit_data.get('CreationTime', ""),
            'user': audit_data.get('MailboxOwnerUPN', ""),
            'ip': audit_data.get('ClientIP', ""),
        }

        for operation in audit_data['OperationProperties']:
            if operation['Name'] == "RuleOperation":
                export['operation'] = operation['Value']
            if operation['Name'] == "RuleId":
                export['rule_id'] = operation['Value']
            if operation['Name'] == "RuleCondition":
                export['rule_condition'] = operation['Value']
            if operation['Name'] == "RuleActions":
                export['rule_actions'] = operation['Value']
            if operation['Name'] == "RuleName" and 'Value' in operation:
                export['rule_name'] = operation['Value']
            else:
                export['rule_name'] = ""
        self.write_to_worksheet('rules', export)
        self.increase_counter('rules')

    # Identify and log New-InboxRule and Remove-InboxRule events
    def analyze_mail_rule(self, audit_data):
        # if you make any changes to the structure of the export field which is written to Excel,
        # those changes must also be made to "analyze_update_inbox_rule" function above
        print(json.dumps(audit_data, indent=4))

        for param in audit_data['Parameters']:
            # if Name in the Parameters, then this is a New Inbox Rule
            if audit_data['Operation'] == "New-InboxRule" and param['Name'] == "Name":
                export = {
                    'datetime': audit_data.get('CreationTime', ""),
                    'user': audit_data.get('UserId', ""),
                    'ip': audit_data.get('ClientIP', ""),
                    'operation': audit_data.get('Operation', ""),
                    'rule_name': param.get('Value'),
                    'rule_id': "",
                    'rule_condition': "",
                    'rule_actions': ''
                }
                self.write_to_worksheet('rules', export)
                self.increase_counter('rules')
            elif audit_data['Operation'] == "Remove-InboxRule" and param['Name'] == "AlwaysDeleteOutlookRulesBlob":
                export = {
                    'datetime': audit_data['CreationTime'],
                    'user': audit_data['UserId'],
                    'ip': audit_data['ClientIP'],
                    'operation': audit_data['Operation'],
                    'rule_name': "",
                    'rule_id': "",
                    'rule_condition': "",
                    'rule_actions': ''
                }
                self.write_to_worksheet('rules', export)
                self.increase_counter('rules')

    # Parses MailItemsAccessed events, reporting individual messages accessed and pulling metadata from Graph
    def analyze_mail_access(self, audit_data):
        for folder in audit_data['Folders']:
            for item in folder['FolderItems']:
                # If clientIP is in the IP ignore list, ignore and move to next
                if audit_data['ClientIPAddress'] in self.ip_ignore_list:
                    continue

                export = {
                    'CreationDate': audit_data['CreationTime'],
                    'UserId': audit_data['UserId'],
                    'Operation': audit_data['Operation'],
                    'Link': "",
                    'Date Received': None,
                    'Sender': "",
                    'Sender Name': "",
                    'Subject': "",
                    'Folder': "",
                    'ClientIP': audit_data['ClientIPAddress'],
                    'MailClient': audit_data['ClientInfoString'],
                    'Throttled': "",
                    'InternetMessageId': item['InternetMessageId']
                }

                for property in audit_data['OperationProperties']:
                    if property['Name'] == "MailAccessType" and property['Name'] == "Bind":
                        export['MailAccessType'] = "Bind"

                    if property['Name'] == "IsThrottled":
                        export['Throttled'] = property['Name']

                # retrieve message metadata from Microsoft
                try:
                    message = self.get_message(export['UserId'], export['InternetMessageId'])
                except Exception as e:
                    print(f"{export['InternetMessageId']} - {e}\n")
                    continue

                # If data in any of these fields is missing, then the message was deleted long enough ago that metadata cannot be retrieved. Auditlog will still reflect the internet message id that was retrieved but nothing more.
                # retrieve message metadata from Microsoft
                try:
                    date = message['value'][0]['receivedDateTime']
                except Exception as e:
                    date = ""

                try:
                    sender = message['value'][0]['from']['emailAddress']['address']
                except Exception as e:
                    sender = ""

                try:
                    sender_name = message['value'][0]['from']['emailAddress']['name']
                except Exception as e:
                    sender_name = ""

                try:
                    mail_folder = audit_data['Folders'][0]['Path']
                except Exception as e:
                    mail_folder = ""

                try:
                    subject = message['value'][0]['subject']
                except Exception as e:
                    subject = ""

                try:
                    link = message['value'][0]['webLink']
                except Exception as e:
                    link = ""

                # update the export dictionary and write line to file
                export.update({
                    'Folder': mail_folder,
                    'Date Received': date,
                    'Sender': sender,
                    'Sender Name': sender_name,
                    'Subject': subject,
                    'Link': link
                })
                self.write_to_worksheet('mail-reads', export)
                self.increase_counter('mail-items')

    # Analyze Mail Sync events - the occur against mail folders and can mean that the entirety
    # of the folder was accessed
    def analyze_mail_sync(self, audit_data):
        export = {}
        if 'ClientProcessName' in audit_data:
            export['user_app'] = audit_data['ClientProcessName']
        else:
            export['user_app'] = ""
        export['datetime'] = audit_data['CreationTime']
        export['operation'] = audit_data['Operation']
        export['user'] = audit_data['UserId']
        export['mailbox'] = audit_data['MailboxOwnerUPN']
        export['user_ip'] = audit_data['ClientIP']
        export['user_agent'] = audit_data['ClientInfoString']

        for prop in audit_data['OperationProperties']:
            if prop['Name'] == "MailAccessType":
                export['access'] = prop['Value']
            if prop['Name'] == "IsThrottled":
                export['throttled'] = prop['Value']

        export['mail_folder'] = audit_data['Item']['ParentFolder']['Name']
        export['mail_folder_path'] = audit_data['Item']['ParentFolder']['Path']

        self.write_to_worksheet('mail-syncs', export)
        self.increase_counter('mail-syncs')

    # same process for deleted messages, will retrive metadata unless the message in unrecoverable
    def analyze_mail_trashed(self, audit_data):
        for item in audit_data['AffectedItems']:
            export = {
                'CreationDate': audit_data['CreationTime'],
                'UserId': audit_data['UserId'],
                'Operation': audit_data['Operation'],
                'Link': "",
                'Date Received': "",
                'Sender': "",
                'Sender Name': "",
                'Subject': "",
                'Folder': "",
                'ClientIP': audit_data['ClientIPAddress'],
                'MailClient': audit_data['ClientInfoString'],
                'InternetMessageId': ""
            }

            # If clientIP is in the IP ignore list, ignore and move to next
            if audit_data['ClientIPAddress'] in self.ip_ignore_list:
                continue

            try:
                internet_message_id = item['InternetMessageId']
            except Exception:
                internet_message_id = ""

            # retrieve message metadata from Microsoft
            try:
                message = self.get_message(export['UserId'], internet_message_id)
            except Exception as e:
                print(f"{internet_message_id} - {e}\n")
                continue

            # If data in any of these fields is missing, then the message was deleted long enough ago that metadata cannot be retrieved. Auditlog will still reflect the internet message id that was retrieved but nothing more.
            try:
                date = message['value'][0]['receivedDateTime']
            except Exception as e:
                date = ""

            try:
                sender = message['value'][0]['from']['emailAddress']['address']
            except Exception as e:
                sender = ""

            try:
                sender_name = message['value'][0]['from']['emailAddress']['name']
            except Exception as e:
                sender_name = ""

            try:
                mail_folder = audit_data['Folders'][0]['Path']
            except Exception as e:
                mail_folder = ""

            try:
                subject = message['value'][0]['subject']
            except Exception as e:
                subject = ""

            try:
                link = message['value'][0]['webLink']
            except Exception as e:
                link = ""

            # update the export dictionary and write line to file
            export.update({
                'Folder': mail_folder,
                'InternetMessageId': internet_message_id,
                'Date Received': date,
                'Sender': sender,
                'Sender Name': sender_name,
                'Subject': subject,
                'Link': link
            })

            self.write_to_worksheet('mail-trashed', export)
            self.increase_counter('mail-trashed')

    # Analyze soft deleted and hard deleted messages
    # soft and hard deleted messages appear have additional metadata to retrieve
    def analyze_deleted_mail(self, audit_data):
        if 'AffectedItems' in audit_data:
            for item in audit_data['AffectedItems']:
                if 'InternetMessageId' in item:
                    imid = item['InternetMessageId']
                else:
                    imid = ""

                if 'Subject' in item:
                    subj = item['Subject']
                else:
                    subj = ""

                export = {
                    'CreationDate': audit_data['CreationTime'],
                    'UserId': audit_data['UserId'],
                    'MailboxOwner': audit_data['MailboxOwnerUPN'],
                    'Operation': audit_data['Operation'],
                    'ClientIP': audit_data['ClientIPAddress'],
                    'MailClient': audit_data['ClientInfoString'],
                    'InternetMessageId': imid,
                    'ParentFolder': item['ParentFolder']['Path'],
                    'Subject': subj,
                }
                self.write_to_worksheet('mail-deleted', export)
                self.increase_counter('mail-deleted')

    # Review messages sent
    def analyze_mail_send(self, audit_data):
        export = {
            'CreationDate': audit_data['CreationTime'],
            'UserId': audit_data['UserId'],
            'Operation': audit_data['Operation'],
            'Link': "",
            'Date Sent': None,
            'Sender': "",
            'Sender Name': "",
            "Subject": "",
            "Folder": "",
            'ClientIP': audit_data['ClientIPAddress'],
            'MailClient': audit_data['ClientInfoString'],
            'InternetMessageId': audit_data['Item']['InternetMessageId']
        }

        # retrieve message metadata from Microsoft
        message = self.get_message(export['UserId'], export['InternetMessageId'])

        # If data in any of these fields is missing, then the message was deleted long enough ago that metadata cannot be retrieved. Auditlog will still reflect the internet message id that was retrieved but nothing more.
        try:
            date = message['value'][0]['sentDateTime']
        except Exception as e:
            date = ""

        try:
            sender = message['value'][0]['from']['emailAddress']['address']
        except Exception as e:
            sender = ""

        try:
            sender_name = message['value'][0]['from']['emailAddress']['name']
        except Exception as e:
            sender_name = ""

        try:
            mail_folder = audit_data['Folders']['Path']
        except Exception as e:
            mail_folder = ""

        try:
            subject = message['value'][0]['subject']
        except Exception as e:
            subject = ""

        try:
            link = message['value'][0]['webLink']
        except Exception as e:
            link = ""

        # update the export dictionary and write line to file
        export.update({
            'Folder': mail_folder,
            'Date Sent': date,
            'Sender': sender,
            'Sender Name': sender_name,
            'Subject': subject,
            'Link': link
        })
        self.write_to_worksheet('mail-sends', export)
        self.increase_counter('mail-sends')

    def analyze_combined_file_operations(self, audit_data):
        export = {
            'date': audit_data.get('CreationTime', ""),
            'operation': audit_data.get('Operation', ""),
            'user': audit_data.get('UserId', ""),
            'file_name': audit_data.get('SourceFileName', ""),
            'full_url': "",
            'client_ip': audit_data.get('ClientIP', ""),
            'user_agent': audit_data.get('UserAgent', ""),
            'platform': audit_data.get('Platform', ""),
            'app_used': "",
            'auth_type': audit_data.get('AuthenticationType', ""),
            'managed_device': audit_data.get('IsManagedDevice', "")
        }

        # Try to extract optional fields
        try:
            export['app_used'] = audit_data['AppAccessContext']['ClientAppName']
        except Exception:
            export['app_used'] = ""

        try:
            export[
                'full_url'] = f"{audit_data['SiteUrl']}/{audit_data['SourceRelativeUrl']}/{audit_data['SourceFileName']}"
        except Exception:
            export['full_url'] = ""

        export['file_name'] = audit_data.get('SourceFileName', "") or audit_data.get('ObjectId', "").split('/')[-1]
        try:
            if audit_data.get('ItemType') == "Folder":
                export['full_url'] = f"{audit_data['SiteUrl']}/{audit_data.get('SourceRelativeUrl', '')}"
            else:
                export['full_url'] = f"{audit_data['SiteUrl']}/{audit_data.get('SourceRelativeUrl', '')}/{audit_data.get('SourceFileName', '')}"
        except Exception:
            export['full_url'] = ""

        self.write_to_worksheet('file-operations', export)
        self.increase_counter('file-operations')

    # Less useful than threat hunting queries, but data isn't subject to expiration as quickly
    # as Threat Hunting Data is
    def analyze_login_events(self, audit_data):
        export = {
            'date': audit_data.get('CreationTime', ""),
            'operation': audit_data['Operation'],
            'workload': audit_data.get('Workload', ""),
            'username': audit_data.get('UserId', ""),
            'ip': audit_data.get('ClientIP', ""),
            'status': "",
            'user_agent': "",
            'result': "",
            'device_name': "",
            'device_os': "",
            "is_compliant": "",
            'error': audit_data.get('LogonError', "").replace("null", "")
        }

        for prop in audit_data['ExtendedProperties']:
            if prop['Name'] == "ResultStatusDetail":
                export['status'] = prop['Value']
            elif prop['Name'] == "UserAgent":
                export['user_agent'] = prop['Value']
            if prop['Name'] == "RequestType":
                export['result'] = prop['Value']

        for prop in audit_data['DeviceProperties']:
            if prop['Name'] == "DisplayName":
                export['device_name'] = prop['Value']
            elif prop['Name'] == "OS":
                export['device_os'] = prop['Value']
            elif prop['Name'] == "IsCompliant":
                export['is_compliant'] = True

        self.write_to_worksheet('logins', export)
        self.increase_counter('logins')

    # Same as Login Events above - Threat Hunting queries are the preferred method of reviewing
    # sign-in history
    def analyze_signin_events(self, audit_data):
        export = {
            'date': audit_data.get('CreationTime', ""),
            'operation': audit_data['Operation'],
            'workload': audit_data.get('Workload', ""),
            'username': audit_data.get('UserId', ""),
            'ip': audit_data.get('ClientIp', ""),
            'status': "",
            'user_agent': audit_data.get('UserAgent'),
            'result': "",
            'device_name': "",
            'device_os': audit_data.get("Platform"),
        }
        self.write_to_worksheet('logins', export)
        self.increase_counter('logins')

    def analyze_sharing_events(self, audit_data):
        export = {
            'timestamp': audit_data.get('CreationTime', ""),
            'operation': audit_data.get('Operation', ""),
            'user': audit_data.get('UserId', ""),
            'user_ip': audit_data.get('ClientIP', ""),
            'target_user': audit_data.get('TargetUserOrGroupName', ""),
            'scope': audit_data.get('SharingLinkScope', ""),
            'permission': audit_data.get('Permission', ""),
            'url': audit_data.get('ObjectId', ""),
            'owner': audit_data.get('SiteUrl', "").split("/")[-1],
            'item_name': audit_data.get('SourceRelativeUrl', ""),
            'user_agent': audit_data.get('UserAgent', ""),
            'browser': audit_data.get('BrowserName', ""),
            'browser_version': audit_data.get('BrowserVersion', ""),
            'application': audit_data.get('ApplicationDisplayName', ""),
            'managed_device': audit_data.get('IsManagedDevice', "")
        }

        sheet_name = "link-activity"

        self.write_to_worksheet(sheet_name, export)
        self.increase_counter(sheet_name)

    # Does the following:
    # * Changes font size to 13pt to be more legible,
    # * widens columns to fit data
    # * Creates a header row and freezes that row to top of page
    # * identifies hyperlinks and turns them into clickable links
    def save_and_cleanup_excel_file(self):
        for sheet in self.workbook.sheetnames:
            worksheet = self.workbook[sheet]

            # Make the entire first row bold
            for cell in worksheet[1]:
                cell.font = Font(bold=True, size=13)
                cell.fill = PatternFill(start_color="ededed", end_color="ededed", fill_type="solid")

            # bump up fon font size and change URLs into working hyperlinks
            for row in worksheet:
                for cell in row:
                    cell.font = Font(size=13)
                    # create hyperlinks in cells that look like they are hyperlinks
                    if type(cell.value) is str and cell.value.startswith("https://"):
                        cell.hyperlink = cell.value
                        cell.value = "View on the Web"
                        # Issue: Neither of the methods below make the link LOOK like a hyperlink
                        # cell.font = Font(bold=True, size=13, underline='single', color='0563C1')
                        cell.style = "Hyperlink"
            # Iterate over all columns and adjust their widths
            # https://python-bloggers.com/2023/05/how-to-automatically-adjust-excel-column-widths-in-openpyxl/
            for column in worksheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(cell.value)
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.1
                worksheet.column_dimensions[column_letter].width = adjusted_width

            # Freeze the header row - openpyxl freezes starting at the row and column BEFORE the cell referenced, so
            # this will freeze the first row
            # https://www.codespeedy.com/freeze-header-rows-in-openpyxl-python/
            worksheet.freeze_panes = 'A2'

        self.workbook.save(f"data{os.sep}{Path(self.input).stem}.xlsx")

    def execute(self):
        i = 0
        print(i, end='')
        with open(self.input) as csv_file:
            csv_reader = csv.DictReader(csv_file)
            for row in csv_reader:
                i = i + 1
                print(f"\rCSV row: {i}", end='', flush=True)

                audit_data = json.loads(row['AuditData'])

                # If operation contains "Rule" (as in New-InboxRule, Remove-InboxRule or any other, then it could indicate
                # a threat actor has modified inbox rules. The user should review the created/changed rules and delete
                # if it's malicious
                if "New-InboxRule" in row['Operation'] or "Remove-InboxRule" in row['Operation']:
                    self.analyze_mail_rule(audit_data)

                if "UpdateInboxRule" in row['Operation']:
                    self.analyze_update_inbox_rule(audit_data)

                # * Sync means the entire mailbox was synced, rather than just batches of messages being retrieved.
                # * Bind just means one or more messages retrieved
                # * Throttled means that logging was throttled - more messages were retrieved than the logs show
                elif row['Operation'] == "MailItemsAccessed" and "Folders" in row['AuditData']:
                    self.analyze_mail_access(audit_data)

                # compile Sent Messages in the same way
                elif row['Operation'] == "Send":
                    self.analyze_mail_send(audit_data)

                # do the same for any messages moved to deleted items
                elif row['Operation'] == "MoveToDeletedItems":
                    self.analyze_mail_trashed(audit_data)

                # Mail folder syncs - much more worrisome
                elif row['Operation'] == "MailItemsAccessed" and "Folders" not in row['AuditData']:
                    self.analyze_mail_sync(audit_data)

                # report soft deleted and hard delete messages. No metadata for these types of messages, but subject
                # is still accessible.
                elif audit_data['Operation'] == "SoftDelete" or audit_data['Operation'] == "HardDelete":
                    self.analyze_deleted_mail(audit_data)

                # OneDrive and Sharepoint Operations
                elif row['Operation'] in [
                    "FileAccessed", "FileAccessedExtended", "FileCheckedIn", "FileCopied",
                    "FileDownloaded", "FileModified", "FileModifiedExtended", "FileMoved",
                    "FilePreviewed", "FileRecycled", "FileRenamed", "FileSyncDownloadedFull",
                    "FileSyncUploadedFull", "FileTranscriptContentAccessed", "FileUploaded"
                ]:
                    self.analyze_combined_file_operations(audit_data)

                # Link sharing events
                if row['Operation'] in [
                    "AddedToSecureLink", "AddedToSharingLink", "AnonymousLinkUsed", "CompanyLinkCreated",
                    "CompanyLinkUsed", "RemovedFromSecureLink", "RemovedFromSharingLink", "SecureLinkCreated",
                    "SecureLinkDeleted", "SecureLinkUpdated", "SecureLinkUsed", "SharingInheritanceBroken",
                    "SharingLinkCreated", "SharingLinkDeleted", "SharingLinkUpdated", "SharingLinkUsed",
                    "SharingPolicyChanged", "SharingRevoked"
                ]:
                    self.analyze_sharing_events(audit_data)

                elif row['Operation'] in [
                    "FolderCreated", "FolderModified", "FolderRenamed"
                ]:
                    self.analyze_combined_file_operations(audit_data)

                elif row['Operation'] in [
                    "UserLoggedIn", "UserLoginFailed"
                ]:
                    self.analyze_login_events(audit_data)

                elif row['Operation'] == "SignInEvent":
                    self.analyze_signin_events(audit_data)

        self.save_and_cleanup_excel_file()


    def main(self):
        args = argparse.ArgumentParser()
        args.add_argument('--ops',
                            default=False,
                            action="store_true",
                            help="Adding this flag just scans the audit file and outputs the names of operations it found")
        args.add_argument('input_file',
                            nargs='?',
                            default=None,
                            help="The path to the audit export file (csv) that you want to process.")
        arg = args.parse_args()
        self.input = Path(arg.input_file)

        if arg.ops == True and arg.input_file != None:
            self.show_operations()

        elif arg.ops == False and arg.input_file != None:
            self.input = arg.input_file
            self.execute()

if __name__ == "__main__":
    analyze = AnalyzeMicrosoftAuditLog()
    analyze.main()

    # analyze.counter dict is not returned as an alphabetical list - this creates a new "report" where the
    # counter names are sorted
    report = {}
    for item in sorted(list(analyze.counter)):
        report[item] = analyze.counter[item]
    report['duration'] = f"{ round(time.time() - analyze.start_time, 2) } seconds"

    print(json.dumps(report, indent=4))